"""
Präferenzkalkulations-Tool für Zollabwicklung
Automatisiert die Berechnung des EU-Präferenzursprungs
Autor: Luis Schenk
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os
import sys

def hauptprogramm():
    """
    Hauptfunktion: Liest Input.xlsx, berechnet Präferenzursprung, schreibt Output.xlsx
    """
    
    print("=" * 50)
    print("Präferenzkalkulations-Tool für Zollabwicklung")
    print("=" * 50)
    
    # 1. PRÜFEN OB INPUT-DATEI EXISTIERT
    input_datei = "Input.xlsx"
    
    if not os.path.exists(input_datei):
        print(f"❌ FEHLER: {input_datei} nicht gefunden!")
        print(f"   Bitte stelle sicher, dass {input_datei} im selben Ordner liegt.")
        sys.exit(1)
    
    # 2. EXCEL-DATEI EINLESEN
    print(f"\n📂 Lese {input_datei}...")
    
    try:
        df = pd.read_excel(input_datei)
    except Exception as e:
        print(f"❌ FEHLER beim Einlesen: {e}")
        sys.exit(1)
    
    # 3. SPALTEN PRÜFEN
    erforderliche_spalten = ['Warennummer', 'Lieferland', 'Materialkosten_Gesamt', 'EU_Anteil']
    fehlende_spalten = [spalte for spalte in erforderliche_spalten if spalte not in df.columns]
    
    if fehlende_spalten:
        print(f"❌ FEHLER: Folgende Spalten fehlen: {', '.join(fehlende_spalten)}")
        print(f"   Vorhandene Spalten: {', '.join(df.columns)}")
        sys.exit(1)
    
    anzahl_zeilen = len(df)
    print(f"✓ {anzahl_zeilen} Zeilen gefunden")
    
    # 4. BERECHNUNGEN DURCHFÜHREN
    print(f"\n🔢 Berechne Präferenzursprung...")
    
    # EU-Anteil in Prozent berechnen
    df['EU_Anteil_Prozent'] = (df['EU_Anteil'] / df['Materialkosten_Gesamt'] * 100).round(2)
    
    # Präferenzursprung bestimmen (JA wenn >= 50%, sonst NEIN)
    df['Präferenzursprung'] = df['EU_Anteil_Prozent'].apply(
        lambda x: 'JA' if x >= 50 else 'NEIN'
    )
    
    # Status-Text hinzufügen
    df['Status'] = df['Präferenzursprung'].apply(
        lambda x: 'Präferenzberechtigt' if x == 'JA' else 'Nicht präferenzberechtigt'
    )
    
    # 5. OUTPUT-DATEI ERSTELLEN
    output_datei = "Output.xlsx"
    print(f"\n💾 Erstelle {output_datei}...")
    
    try:
        df.to_excel(output_datei, index=False)
    except Exception as e:
        print(f"❌ FEHLER beim Schreiben: {e}")
        sys.exit(1)
    
    # 6. FORMATIERUNG HINZUFÜGEN (Grün für JA, Rot für NEIN)
    print(f"🎨 Formatiere Ergebnisse...")
    
    try:
        # Excel-Datei mit openpyxl öffnen
        wb = load_workbook(output_datei)
        ws = wb.active
        
        # Farben definieren
        gruen = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Hellgrün
        rot = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")     # Hellrot
        
        # Spaltenindex für Präferenzursprung finden (normalerweise Spalte F)
        praefrenz_spalte = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == 'Präferenzursprung':
                praefrenz_spalte = idx
                break
        
        if praefrenz_spalte:
            # Durch alle Zeilen gehen (ab Zeile 2, da Zeile 1 = Header)
            for zeile in range(2, ws.max_row + 1):
                wert = ws.cell(row=zeile, column=praefrenz_spalte).value
                
                # Ganze Zeile färben
                if wert == 'JA':
                    for spalte in range(1, ws.max_column + 1):
                        ws.cell(row=zeile, column=spalte).fill = gruen
                elif wert == 'NEIN':
                    for spalte in range(1, ws.max_column + 1):
                        ws.cell(row=zeile, column=spalte).fill = rot
        
        # Speichern
        wb.save(output_datei)
        
    except Exception as e:
        print(f"⚠️  Warnung: Formatierung fehlgeschlagen: {e}")
        print(f"   Die Datei wurde trotzdem erstellt (ohne Farben).")
    
    # 7. ZUSAMMENFASSUNG
    ja_anzahl = len(df[df['Präferenzursprung'] == 'JA'])
    nein_anzahl = len(df[df['Präferenzursprung'] == 'NEIN'])
    
    print("\n" + "=" * 50)
    print("✅ FERTIG!")
    print("=" * 50)
    print(f"📊 Ergebnis:")
    print(f"   • Präferenzberechtigt (JA):    {ja_anzahl} von {anzahl_zeilen}")
    print(f"   • Nicht präferenzberechtigt (NEIN): {nein_anzahl} von {anzahl_zeilen}")
    print(f"\n💾 {output_datei} wurde erstellt!")
    print("=" * 50)

# PROGRAMM STARTEN
if __name__ == "__main__":
    hauptprogramm()